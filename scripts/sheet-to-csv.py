#!/usr/bin/env python3
"""Pull one tab out of the workbook as CSV, for import-sheet.mjs.

    python3 scripts/sheet-to-csv.py <file.xlsx> [tab] > out.csv

Exists because the round trip through Google Sheets mangles two things that
then confuse the importer, and both are fixed here rather than in fourteen
places downstream.

INTEGERS COME BACK AS FLOATS. Sheets stores every number as a double, so a
pack count of 9 exports as "9.0" and a view count of 759 as "759.0". Anything
whole is written back as a plain integer.

FORMULA CELLS ARE READ AS VALUES, not as "=SUM(H5,J5,L5,N5,P5)". Packs Opened
is a computed column, and the importer wants the number it computed.

Trailing spaces are stripped too. Sheets autocomplete leaves them behind
("Surging Sparks ") and a set name with a space on the end matches nothing.
"""
import csv
import sys
from pathlib import Path

from openpyxl import load_workbook

src = Path(sys.argv[1]) if len(sys.argv) > 1 else None
tab = sys.argv[2] if len(sys.argv) > 2 else "Video Log"
if not src or not src.exists():
    sys.exit(f"usage: sheet-to-csv.py <file.xlsx> [tab]\nnot found: {src}")

wb = load_workbook(src, data_only=True)
if tab not in wb.sheetnames:
    sys.exit(f'no tab named "{tab}". Tabs: {", ".join(wb.sheetnames)}')
ws = wb[tab]

# A FORMULA WITH NO CACHED VALUE READS AS BLANK, AND THAT IS HOW 1,062 PACKS
# WENT TO ZERO. data_only=True asks for the value Excel last computed and
# stored; openpyxl never writes one, so every formula in a file THIS PROJECT
# generated comes back None. Packs Opened is handled downstream by adding the
# per-set cells instead, but any other formula -- one typed into a cell by
# hand, a stray =SUM left over from working something out -- disappears with
# nothing said. So the same file is opened a second time without data_only,
# and every formula cell that came back empty is named on stderr. stderr,
# because stdout is the CSV.
_wf = load_workbook(src, data_only=False)[tab]
_lost = []

out = csv.writer(sys.stdout)
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
    cells = []
    for c in row:
        v = c.value
        if v is None:
            f = _wf.cell(c.row, c.column).value
            if isinstance(f, str) and f.startswith("="):
                _lost.append(f"{c.coordinate}  {f[:60]}")
            cells.append("")
        elif isinstance(v, float) and v.is_integer():
            cells.append(str(int(v)))
        else:
            cells.append(str(v).strip())
    out.writerow(cells)

# Two formulas are on every row of the Video Log by design and neither is read
# as a value: Packs Opened is recomputed from its parts by the importer, and
# Watch is a convenience link nothing imports. Counted, not listed, so the
# warning only ever fires on a formula somebody put there.
_ROUTINE = ("=SUM(", "=HYPERLINK(")
_expected = sum(1 for x in _lost if x.split()[1].startswith(_ROUTINE))
_other = [x for x in _lost if not x.split()[1].startswith(_ROUTINE)]
if _other:
    print(f"\n{len(_other)} formula cell(s) on '{tab}' had no stored value and were read as BLANK.",
          file=sys.stderr)
    print("  Excel and Google Sheets store one; a file written by this project does not.",
          file=sys.stderr)
    print("  Open the sheet, replace the formula with its answer, and export again.",
          file=sys.stderr)
    for x in _other[:15]:
        print("  " + x, file=sys.stderr)
elif _expected:
    print(f"({_expected} computed Packs Opened cells read as blank, as expected; "
          "the per-set Packs columns are the source.)", file=sys.stderr)
