#!/usr/bin/env python3
"""Build the video log workbook for Google Sheets.

    python3 scripts/build-sheet.py

Writes Garbage-Rips-585-Video-Log.xlsx at the repo root, prefilled from
public/data/videos.json and sets.json. Every column the site actually reads is
here, and every column is either something YouTube already knows (grey, do not
edit) or something only Tim knows (yellow, please fill in).

Where a tag was derived automatically the guess is prefilled, so the job is
correcting rather than typing. scripts/import-sheet.mjs reads the CSV export
back and turns it into data/overrides.json and data/manual.json.

Safe to re-run: it rebuilds from scratch, so run import-sheet.mjs on your
latest export BEFORE rebuilding, or your edits are what gets overwritten.
"""
import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "Garbage-Rips-585-Video-Log.xlsx"

# What has already been imported from a filled-in sheet. Restored onto the rows
# below so a rebuild returns a human's answers rather than overwriting them.
try:
    manual = json.loads((ROOT / "data/manual.json").read_text())
except Exception:
    manual = {}

# Only one thing is read out of here: an explicit empty set list, which is the
# stored form of "Not a set (sealed/other)" and is indistinguishable from "not
# answered" by the time it reaches videos.json.
try:
    overrides_src = {k: v for k, v in json.loads((ROOT / "data/overrides.json").read_text()).items()
                     if isinstance(v, dict)}
except Exception:
    overrides_src = {}

videos = json.loads((ROOT / "public/data/videos.json").read_text())
videos = videos.get("videos", videos)

# ---------------------------------------------------------------------------
# WHAT TIM'S OWN TITLE ALREADY SAID, read back so he does not type it twice.
# ---------------------------------------------------------------------------
#
# Asked for by name: "we should like videos by what they are, example, Chaos
# Rising ETB 3 - Pack 3, or Pitch Black Booster Bundle 2 Pack 6, so that way we
# can get the accurate data on how many of each product type we opened overall
# and for each set, and then also so we know exactly how many packs we have
# ripped overall for each set and overall."
#
# 235 of 316 videos state a pack number and 67 state a box number, most of them
# in the title, in his own words: "Mega Zygarde Box #2 Pack #5". Handing back a
# blank cell for a fact the title spells out is 300 rows of retyping.
#
# TRANSCRIPTION, NOT INFERENCE, AND THE LINE MATTERS. The comment further down
# this file explains why a guessed HIT was removed: a colour does not survive
# export to CSV, so the importer read every blue suggestion back as a typed
# answer, and 62 of 64 rarities in the log arrived that way. The defence here is
# not the colour, it is the pattern. These two regexes match ONLY the literal
# words "Box"/"ETB"/"Bundle" or "Pack" followed by a number. They do not read
# ordinal prose, so "our third Chaos Rising ETB" is deliberately NOT matched
# even though a person can see it means 3: that is the sentence the older
# comment below warns about, because a description can say it twice with two
# different numbers ("Pack #2 of our third ETB") and a matcher picks wrong
# confidently. A number Tim typed after the word Box is his answer being read
# back. A number a rule worked out from prose is the machine's guess, and this
# file does not put those in the same colour, it does not put them in at all.
try:
    _desc = json.loads((ROOT / "data/descriptions.json").read_text())
except Exception:
    _desc = {}

# "Box #2", "ETB #3", "Bundle 2". The word first, then the number.
BOX_RE = re.compile(r"\b(?:box|etb|elite\s+trainer\s+box|booster\s+bundle|bundle|tin|upc)\s*#?\s*(\d{1,2})\b", re.I)
# "Pack #5", "Pack 5". A bare "#5" is NOT matched: on this channel it is as
# often a card number or a set number as a pack.
PACK_RE = re.compile(r"\bpack\s*#?\s*(\d{1,2})\b", re.I)


def _stated(v, rx):
    """The number this video's own title or description states, or None.

    Title first: when it is there it is the headline fact and it is the copy
    Tim wrote most deliberately. Description second, because that is where most
    of them actually live.
    """
    for s in (v.get("title") or "", _desc.get(v.get("id"), "") or ""):
        m = rx.search(s)
        if m:
            n = int(m.group(1))
            if 1 <= n <= 40:
                return n
    return None

sets = json.loads((ROOT / "public/data/sets.json").read_text())["sets"]
set_name = {s["id"]: s["name"] for s in sets}

# The non-English guides. Without these the 21 imported rips had no option in
# the Set dropdown at all, and their prefilled guess came back blank, so the one
# part of the catalogue that most needed a human answer was the one part the
# sheet could not ask about. Labels match shared/taxonomy.mjs exactly, because
# import-sheet.mjs maps this text straight back to a set id.
_LANG_TAG = {"ja": "JP", "ko": "KR", "zh-cn": "CN", "zh-tw": "CN"}
intl_sets = {}
try:
    _ig = json.loads((ROOT / "public/data/intl-guides.json").read_text())["sets"]
    for _id, _g in _ig.items():
        intl_sets[_id] = f"{_g['english']} ({_LANG_TAG.get(_g.get('lang'), '??')})"
except Exception:
    pass
set_name.update(intl_sets)

# ---------------------------------------------------------------- vocabulary

# EVERY ENGLISH SET EVER PRINTED, not only the ones with a guide page.
#
# This used to offer the 28 sets that have a guide. Tim's own reason for wanting
# more: a box set or a tin holds packs from sets we have no guide for, and until
# now there was no cell he could fill to record that. He hit it on a Black Bolt
# pack and had nowhere to put the answer.
#
# public/data/expansions.json is the full list from the Pokemon TCG API, 174
# sets, and their names are unique, so a name is a safe key. Newest first,
# because the sets being opened on camera are almost always current and a
# 180-row list is a long scroll to reach this year.
#
# THIS DOES NOT TEACH THE MATCHER 146 NEW NAMES. These are options for Tim to
# pick, not patterns for the automatic tagger to guess from. Widening what the
# matcher guesses at is what retagged "Pitch Black is SAVAGE Today!" as a 2019
# McDonald's set, and tags now come from his answers rather than from title
# matching anyway.
try:
    _exp = json.loads((ROOT / "public/data/expansions.json").read_text())["sets"]
    _guided = {s["name"] for s in sets}
    _rest = [e["name"] for e in sorted(_exp, key=lambda e: e.get("released") or "", reverse=True)
             if e.get("name") and e["name"] not in _guided]
except Exception:
    _rest = []

SET_NAMES = ([s["name"] for s in sets]
             + sorted(intl_sets.values())
             + _rest
             + ["Multiple sets", "Not a set (sealed/other)"])

OPENING_TYPES = [
    "Single Booster Pack", "Booster Bundle", "Booster Box",
    "ETB (Elite Trainer Box)", "SPC (Super Premium Collection)",
    "UPC (Ultra Premium Collection)", "Poke Ball Tin", "Tin",
    "ex Premium Collection", "ex Special Collection", "ex Box", "Collection Box",
    # Added after it was typed in by hand, because the dropdown did not offer it
    # and the product is real. Anything typed freehand into a validated column
    # is a missing option, not a user error.
    "Knock Out Collection",
    "Blister", "Japanese Booster Pack", "Korean Booster Pack",
    "Chinese Booster Pack", "Other",
]

# Mirrors the ladder the site ranks by, in the same order, so the Hall of Fame
# sorts the way the sheet reads.
# THE STAR HINT USED TO BE INSIDE THE VALUE and it broke the column. The options
# read "Hyper Rare (3 gold stars)" while everything this script PREFILLED, and
# everything the site stores, is the bare name "Hyper Rare". So 63 of the 64
# filled cells failed validation the moment the file reached Google Sheets, and
# the column came back covered in warning triangles. One cell matched: the one
# where the long form was picked from the dropdown by hand.
#
# A dropdown value is a KEY. It has to equal what the data says, or the sheet
# argues with itself. The hint is genuinely useful, so it moved to a comment on
# the header cell and to the Read Me, where it costs nothing.
RARITIES = [
    "Mega Hyper Rare",
    "Hyper Rare",
    "Special Illustration Rare",
    "Illustration Rare",
    "Ultra Rare",
    "Double Rare",
    "Rare",
    "ACE SPEC Rare",
    "Super Rare",
    "Charizard",
    # Tim used this on two videos where a single rip produced several notable
    # cards. It was not an option and it should have been: the alternative is
    # forcing one card to stand for the rip. The per-card detail goes on the
    # My Hits tab.
    "More than one",
    "No hit",
]

# The star hint, kept out of the value and shown as a note on the header instead.
RARITY_HINT = (
    "Read the star row at the bottom of the card.\n"
    "Mega Hyper Rare: one big yellow star.\n"
    "Hyper Rare: three gold stars.\n"
    "Special Illustration Rare: two gold stars.\n"
    "Illustration Rare: one gold star.\n"
    "Ultra Rare: two silver stars.\n"
    "Double Rare: two black stars.\n"
    "Rare: one black star.\n"
    "ACE SPEC Rare: one pink star.\n"
    "Charizard: any rarity, it is its own category here.\n"
    "More than one: several notable cards, list them on My Hits."
)

YESNO = ["Yes", "No"]
PLAYLISTS = ["Greatest Hits", "Hits Only", "Full Box Openings", "Singles",
             "Japanese", "None"]

# Which derived pull tag maps onto which sheet rarity, so the guess is prefilled.
PULL_TO_RARITY = {
    "gold": "Hyper Rare",
    "sir": "Special Illustration Rare",
    "ir": "Illustration Rare",
    "double-rare": "Double Rare",
    "charizard": "Charizard",
}
PULL_ORDER = ["gold", "sir", "ir", "double-rare", "charizard"]

# Box and series names, derived from the titles rather than invented, so the
# dropdown matches what Tim actually films. Only names that begin with a set
# name or "Mega" survive: the rest of a title is clickbait, not a product.
_set_words = tuple(x["name"].lower() for x in sets) + ("mega ",)
BOX_NAMES = set()
for _v in videos:
    _head = re.sub(r"Pack\s*#\d+", "", _v.get("title", "").split("|")[0], flags=re.I)
    _head = re.sub(r"#\d+", "", _head)
    _head = re.sub(r"[^\w\s&'.:-]", "", _head)
    _head = re.sub(r"\s+", " ", _head).strip()
    if 6 < len(_head) < 48 and _head.lower().startswith(_set_words) \
       and re.search(r"box|tin|etb|collection|bundle|upc|blister|premium", _head, re.I):
        BOX_NAMES.add(_head)

# SEALED PRODUCTS THAT ARE NOT EXPANSIONS, and this is where their name lives.
#
# Three rips open something with a real product name and no set behind it:
# a Trick or Trade bundle (Halloween mini packs), a Victini Illustration
# Collection and a Mega Heroes mini tin. Neither api.pokemontcg.io nor TCGdex
# lists any of the three among their 174 and 218 sets, so they are products,
# not expansions, and putting them in sets.json would mean inventing a card
# count and a release date for something that has neither. They stay out.
#
# But the question "what did he open" still HAS an answer, and without these
# three rows the only honest thing Tim could put in the Set column was
# "Not a set (sealed/other)", which records that there is no expansion and
# throws away which product it was. This column is where the answer survives:
# import-sheet.mjs stores it as `box` in data/manual.json and this script reads
# it straight back, so it round trips like any other typed answer.
#
# The derivation above cannot reach them: it only keeps a title that STARTS
# with a set name or "Mega", and all three of these titles open with "Pokémon"
# or with the channel's own catchphrase. That filter is right and these are the
# exception, so they are listed rather than the filter loosened.
BOX_NAMES.update([
    "Trick or Trade BOOster Bundle",
    "Victini Illustration Collection",
    "Mega Heroes Mini Tin",
])
BOX_NAMES = sorted(BOX_NAMES)

# Where a graded price came from. Naming the source is what makes the number
# checkable later.
PSA_SOURCES = ["PSA Price Guide", "eBay sold listings", "130point", "TCGplayer",
               "PriceCharting", "Local shop", "Other"]

SHOP_AREAS = ["Rochester, NY", "Greece, NY", "Henrietta, NY", "Irondequoit, NY",
              "Brighton, NY", "Webster, NY", "Pittsford, NY", "Victor, NY",
              "Buffalo, NY", "Syracuse, NY", "Online only"]

SHOP_GOOD_FOR = ["Sealed product", "Singles", "New releases", "In store",
                 "Graded slabs", "Vintage", "Japanese", "Events and tournaments",
                 "Trade ins", "Supplies"]

HOF_RANKS = [str(i) for i in range(1, 21)]

# WHICH ONE OF THESE, AND WHICH PACK OUT OF IT. Two dropdowns of plain numbers.
#
# They are lists rather than free cells for the same reason every other column
# here has one: it makes the column obviously a NUMBER, and picking 3 is one
# click. Neither is strict, so a fourth Booster Box or the 37th pack of
# something exotic is typed straight in. The ranges are chosen from what the
# products actually hold: nothing on the Opening Type list holds more than a
# booster box's 36 packs, and 30 of one product from one set is far past
# anything the channel has opened.
BOX_NUMBERS = [str(i) for i in range(1, 31)]
PACK_NUMBERS = [str(i) for i in range(1, 37)]

# Every chase card we know about, so Hit Card is a pick rather than a spelling
# test. Not exhaustive on purpose: the validation is a suggestion, not a rule.
def _hit_card_list():
    """Cards worth logging as a hit, dearest first, deduped by name.

    Was the 152 chase cards from sets.json, which is every card we happened to
    show on a set page. The full checklist is now on disk, so this offers the
    cards someone would actually write down: anything over $5, capped so the
    dropdown stays usable in Excel. Still a suggestion, not a rule.
    """
    seen = {}
    try:
        for f in sorted((ROOT / "public/data/cards").glob("*.json")):
            for c in json.loads(f.read_text())["cards"]:
                price = c.get("price")
                if not isinstance(price, (int, float)) or price < 5:
                    continue
                name = c.get("name") or ""
                if name and price > seen.get(name, 0):
                    seen[name] = price
    except Exception:
        pass
    for st in sets:                       # keep the old source as a floor
        for c in (st.get("chase") or []):
            seen.setdefault(c["name"], 0)
    ranked = sorted(seen.items(), key=lambda kv: -kv[1])[:500]
    return sorted(n for n, _ in ranked)

HIT_CARDS = _hit_card_list()

# Every product id shared/taxonomy.mjs can assign, mapped to the sheet's own
# wording. Eight were missing, so a video already tagged booster-box or
# poke-ball-tin came up with a BLANK Opening Type instead of a prefilled guess,
# and the whole correct-the-guess workflow silently skipped box openings.
PRODUCT_TO_OPENING = {
    "single-pack": "Single Booster Pack",
    "bundle": "Booster Bundle",
    "booster-box": "Booster Box",
    "etb": "ETB (Elite Trainer Box)",
    "spc": "SPC (Super Premium Collection)",
    "upc": "UPC (Ultra Premium Collection)",
    "poke-ball-tin": "Poke Ball Tin",
    "tin": "Tin",
    "ex-premium": "ex Premium Collection",
    "ex-special": "ex Special Collection",
    # ex-box USED TO BE ITS OWN ID AND HAD ITS OWN LABEL HERE, because both ids
    # once pointed at the same string and the importer mapped that string back
    # to ex-premium, so every re-import silently converted 24 ex-box videos into
    # ex Premium Collections. shared/taxonomy.mjs has since folded ex-box into
    # ex-premium outright -- all 24 were hand-corrected to it anyway -- so there
    # is no longer an id for this row to prefill and it would be dead code. The
    # dropdown still OFFERS "ex Box", because that is what the titles say, and
    # the importer now reads it as ex-premium.
    "collection-box": "Collection Box",
    # knock-out was added to the taxonomy, to OPENING_TYPES and to the
    # importer's PRODUCT_IDS on the same day and missed here, which is the ninth
    # time this map has been the one that lagged. The effect is the one the
    # comment above describes: a video the matcher already tags knock-out gets a
    # BLANK Opening Type, so the correct-the-guess workflow skips it entirely.
    "knock-out": "Knock Out Collection",
    "blister": "Blister",
    "japanese-pack": "Japanese Booster Pack",
    "korean-pack": "Korean Booster Pack",
    "chinese-pack": "Chinese Booster Pack",
}

# Standard pack counts per product, prefilled in blue as a guess to be
# confirmed. These are the usual contents, not a guarantee: pack counts change
# between eras and between regions, and a "tin" is anything from 3 to 5. The
# ones that genuinely vary are left blank rather than guessed, because a wrong
# denominator is worse on the luck page than a missing one.
PRODUCT_TO_PACKS = {
    "single-pack": 1,
    "japanese-pack": 1,
    "korean-pack": 1,
    "chinese-pack": 1,
    "blister": 3,
    "bundle": 6,
    "etb": 9,
    "booster-box": 36,
    "upc": 16,
    "spc": 8,
}

# ------------------------------------------------------------------- styling

BODY = Font(name="Arial", size=10)
BOLD = Font(name="Arial", size=10, bold=True)
TITLE = Font(name="Arial", size=14, bold=True)
NOTE = Font(name="Arial", size=10, italic=True, color="555555")
LOCKED_TXT = Font(name="Arial", size=10, color="666666")
GUESS_TXT = Font(name="Arial", size=10, color="0000FF")   # auto-filled, confirm

HEAD_LOCKED = PatternFill("solid", fgColor="D9D9D9")   # from YouTube
HEAD_INPUT = PatternFill("solid", fgColor="FFF2CC")    # yours to fill
HEAD_HOF = PatternFill("solid", fgColor="FFD966")      # the Hall of Fame block
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = Workbook()

# =========================================================== 1. Read Me =====

ws = wb.active
ws.title = "Read Me"
ws.column_dimensions["A"].width = 26
ws.column_dimensions["B"].width = 104

rows = [
    ("Garbage Rips 585 master sheet", None, TITLE),
    (None, None, None),
    ("What this is", "The control surface for garbagerips.com. Everything the site knows that "
                     "cannot be worked out from YouTube or the card database lives here. Fill in "
                     "what you know, export, import, and the site changes.", None),
    (None, None, None),
    ("Grey columns", "From YouTube or the card database. Do not edit. Video ID and Key are what "
                     "link a row to its video or card, so never edit those and never sort in a way "
                     "that separates a row from them.", None),
    ("Yellow columns", "Yours. Blank is fine everywhere: the import skips anything unanswered, so "
                       "twenty rows at a time is a perfectly good way to work.", None),
    ("Gold columns", "Hall of Fame and Most Wanted, the two things that change the top of the "
                     "home page.", None),
    ("Blue text", "A guess the site already made. Correct it if it is wrong, leave it if it is "
                  "right. Black text means a person typed it.", None),
    ("Dropdowns", "A shortcut, never a gate. Every one accepts anything you type, so a value "
                  "missing from the list is never a reason you cannot record it.", None),
    ("If the list looks short",
                  "Google Sheets does not keep a dropdown that points at another tab. When you "
                  "export back to xlsx it rebuilds each one from the values already in that "
                  "column, so options nobody has used yet disappear. The Lists tab is the real "
                  "list. Read it there and type the value in; it will be accepted.", None),
    ("Hit Card is free text",
                  "It has no dropdown here. If Sheets offers to make one out of what you have "
                  "typed, decline it. If one appears anyway, select the column and use "
                  "Data then Data validation then Remove rule.", None),
    (None, None, None),
    ("THE TABS", None, BOLD),
    ("Video Log", "Every video. Sets, opening type, the hit, the Hall of Fame, and per-video copy "
                  "for the site.", None),
    ("Set Notes", "The two facts the card database does not carry: whether a set is still in "
                  "print, and what a pack costs. Set guides omit both until filled in.", None),
    ("Chase Cards", "Every chase card on the site. PSA 10 prices live ONLY here: nothing can fetch "
                    "a graded price, so this tab is the only route onto the site. Most Wanted "
                    "marks a card for the hunt page and the home page band.", None),
    ("Shops", "Local card shops for /shops.html.", None),
    ("Summary", "Live counts of how much is filled in.", None),
    ("Lists", "The source of every dropdown. Add a row here to add an option.", None),
    (None, None, None),
    ("HOW TO IMPORT", "Google Sheets exports the ACTIVE tab only, so open the tab you want first, "
                      "then File > Download > Comma-separated values.", BOLD),
    ("Video Log", "node scripts/import-sheet.mjs <csv>", None),
    ("Chase Cards", "node scripts/import-cards.mjs <csv>", None),
    ("Shops", "node scripts/import-cards.mjs <csv>", None),
    (None, "Either importer works out which tab it is from the header row, so the filename does "
           "not matter.", None),
    (None, None, None),
    ("The one that matters most", "Set. 61 videos have no set tag. A video with no set cannot be "
                                  "filtered and cannot reach the Hall of Fame. Eight of those are "
                                  "graded hits locked out of the home page right now.", None),
    (None, None, None),
    ("Packs from several sets", "A tin with two packs, or a box with ten packs from four sets, "
                                "belongs to every set it contains. Put the set the video is really "
                                "about in Set and the rest in Set 2 to Set 5. Next to each one, put "
                                "how many packs of THAT set were in the opening. Packs Opened adds "
                                "them up for you, so a UPC reads 18 packs AND tells us six were "
                                "Journey Together. That is what makes a per-set pack count possible.", None),
    (None, None, None),
    ("Which one, and which pack",
                  "Box # is which one of that product you are opening, counting your own openings "
                  "of it from that set: your third Chaos Rising ETB is 3. Pack # is which pack out "
                  "of it this video opens. Fill both in and the Chaos Rising guide can say how many "
                  "ETBs, bundles and single packs have been opened from that set, which nothing "
                  "else in the sheet can work out. Both are optional and blank is the normal case: "
                  "a set with none recorded shows no counts at all rather than a zero.", None),
    (None, None, None),
    ("Greatest Hits", "Mark Yes on the RIPS you want in the gold section at the top of the home "
                      "page. Rank orders them, 1 first. Leave Rank blank and the site orders by "
                      "rarity then views. These are videos.", None),
    (None, None, None),
    ("Card Hall of Fame", "On the Chase Cards tab. Mark Yes on the CARDS you have actually pulled. "
                          "They get their own page, ranked by value, with the card image, the set, "
                          "the raw near mint price and the PSA 10 price. These are cards, not "
                          "videos, which is the whole difference between the two.", None),
    (None, None, None),
    ("PSA 10 prices", "Put in a number you have actually seen, set PSA 10 Checked to the date you "
                      "checked it (2026-08-11 format), and name the source. A graded price with no "
                      "date is not a fact about anything, and a blank shows nothing rather than a "
                      "guess.", None),
    (None, None, None),
    ("Careful", "Rebuilding this file overwrites it. Import your latest export FIRST, then rebuild. "
                "Anything already imported is carried back in; anything not is lost.", NOTE),
]
for i, (a, b, font) in enumerate(rows, start=1):
    if a is not None:
        ws.cell(i, 1, a).font = font or BOLD
    if b is not None:
        c = ws.cell(i, 2, b)
        c.font = font or BODY
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[i].height = 30 if b and len(str(b)) > 95 else None

# ============================================================= 2. Lists =====

wl = wb.create_sheet("Lists")
for col, (head, items) in enumerate(
    [("Sets", SET_NAMES), ("Opening Types", OPENING_TYPES), ("Hit Rarities", RARITIES),
     ("Yes/No", YESNO), ("Playlists", PLAYLISTS), ("Box / Series", BOX_NAMES or ["(none yet)"]),
     ("Hit Cards", HIT_CARDS or ["(none yet)"]), ("PSA 10 Sources", PSA_SOURCES),
     ("Shop Areas", SHOP_AREAS), ("Shop Good For", SHOP_GOOD_FOR), ("HoF Ranks", HOF_RANKS),
     # APPENDED, NOT INSERTED. named() below addresses these columns by NUMBER,
     # so putting a new list anywhere but the end shifts every dropdown after it
     # onto the wrong column. Same trap the Video Log solved by looking columns
     # up by header; here the fix is simply to only ever add at the end.
     ("Box Numbers", BOX_NUMBERS), ("Pack Numbers", PACK_NUMBERS)],
    start=1,
):
    ws_c = get_column_letter(col)
    wl.column_dimensions[ws_c].width = max(18, len(head) + 4, max(len(i) for i in items) + 2)
    wl.cell(1, col, head).font = BOLD
    for r, item in enumerate(items, start=2):
        wl.cell(r, col, item).font = BODY

# A DEFINED NAME PER LIST, NOT A RAW CROSS-SHEET RANGE.
#
# The dropdowns used to point straight at Lists!$A$2:$A$44. That is correct xlsx
# and Excel honours it, but the sheet does not live in Excel: it goes through
# Google Sheets, and a dropdown that points at another tab does not survive the
# trip intact. Measured on one round trip, the Set column came back offering 35
# of the 43 sets this file writes, having quietly dropped every option no row
# had used yet, Black Bolt and White Flare among them. Tim opened a Black Bolt
# pack, went to record it, and the sheet had no such value.
#
# A workbook-level defined name is the same range with a stable label on it, and
# a label survives being carried between applications far better than a raw
# reference does. The lists still live on the Lists tab and are still the thing
# to read when a dropdown looks short.
#
# THIS IS A BEST EFFORT AND IT IS DELIBERATELY NOT THE ONLY DEFENCE. Nothing
# written here can compel another application to keep a dropdown. That is why
# every validation is also non-strict: if the list arrives short anyway, the
# value can still be typed and the importer reports anything it does not
# recognise rather than swallowing it.
DEFINED = {}

def named(col_idx, count):
    """A workbook-level defined name covering one list column on Lists."""
    c = get_column_letter(col_idx)
    ref = f"Lists!${c}$2:${c}${count + 1}"
    label = f"gr_list_{c}"
    DEFINED[label] = ref
    return label

DV_SET = named(1, len(SET_NAMES))
DV_OPEN = named(2, len(OPENING_TYPES))
DV_RARITY = named(3, len(RARITIES))
DV_YESNO = named(4, len(YESNO))
DV_PLAYLIST = named(5, len(PLAYLISTS))
DV_BOX = named(6, len(BOX_NAMES) or 1)
DV_HITCARD = named(7, len(HIT_CARDS) or 1)
DV_PSASRC = named(8, len(PSA_SOURCES))
DV_AREA = named(9, len(SHOP_AREAS))
DV_GOODFOR = named(10, len(SHOP_GOOD_FOR))
DV_RANK = named(11, len(HOF_RANKS))
DV_BOXNO = named(12, len(BOX_NUMBERS))
DV_PACKNO = named(13, len(PACK_NUMBERS))


def dv(formula, strict=False):
    """A dropdown that offers the list and still accepts anything typed.

    STRICT USED TO BE THE DEFAULT AND IT LOCKED TIM OUT OF HIS OWN SHEET.

    Google Sheets does not keep a dropdown that points at a range on another
    tab. On export back to xlsx it rewrites each one as a literal list built
    from THE VALUES ALREADY IN THAT COLUMN, so an option nobody has used yet
    simply disappears. Measured on one round trip: the Set column went from the
    43 options this file writes to 35, losing Black Bolt, White Flare, Stellar
    Crown, Shrouded Fable, Paldean Fates and five more. Tim opened a Black Bolt
    pack, went to record it, and the sheet refused the value.

    Strict validation turns that from a missing convenience into a wall. Not
    strict, the dropdown is a shortcut when the list is right and simply gets
    typed past when it is not, and the importer still reports anything it does
    not recognise rather than swallowing it.

    The Lists tab stays the authoritative list and is the thing to read when the
    dropdown looks short. Nothing here can stop Sheets rewriting it."""
    return DataValidation(type="list", formula1=formula, allow_blank=True,
                          showDropDown=False, showErrorMessage=strict)

# ========================================================= 3. Video Log =====

wv = wb.create_sheet("Video Log")

COLUMNS = [
    # (header, width, kind)  kind: locked | input | hof
    ("Video ID", 14, "locked"),
    ("Title", 52, "locked"),
    ("Published", 12, "locked"),
    ("Views", 9, "locked"),
    ("Length", 8, "locked"),
    ("Watch", 8, "locked"),
    # SET AND ITS PACK COUNT SIT TOGETHER, five pairs of them.
    #
    # A UPC is 18 packs across five sets. Recording only the total meant we knew
    # 18 packs were opened but not that six were Journey Together, so no per-set
    # pack count could ever be computed and "how many Pitch Black packs have we
    # opened" had no answer. The pair is the smallest thing that fixes it.
    #
    # FIVE PAIRS, not a separate tab, because the distribution says so: of 272
    # videos with any set recorded, 259 are a SINGLE set, 11 are two, one is
    # three and one is five. A tab would make 259 rows a two-place job to serve
    # thirteen. Five pairs covers every video in the catalogue today.
    #
    # This replaces "More Sets", which was free text. Free text in one cell is
    # exactly what broke the hits: one typo and one stray comma cost two cards.
    # ORDER FOLLOWS THE ORDER A PERSON ANSWERS THE QUESTIONS, which is not the
    # order the data model lists them in. Watching a rip you know WHAT WAS
    # OPENED before you know which sets were inside it: Opening Type is read off
    # the box in the first two seconds, and the set is often only certain once a
    # pack is out. Opening Type also drives what Box # and Pack # even mean, so
    # asking for it first means the next three columns arrive already framed.
    ("Opening Type", 28, "input"),
    ("Set", 24, "input"),
    ("Packs", 8, "input"),
    ("Box / Series", 46, "input"),
    # WHICH ONE OF THESE, AND WHICH PACK OUT OF IT.
    #
    # Asked for by name: "we should be able to see I have opened 3 Chaos Rising
    # ETBs and this is pack 3 from that ETB". Opening Type says an ETB was
    # opened and Packs says how many packs the ETB holds. Neither says which ETB
    # or which pack, and without that the log cannot answer "how many ETBs of
    # this set have we been through", which is the question.
    #
    # THEY SIT IN THE DAILY BLOCK ON PURPOSE, right after the column they
    # qualify. The four extra Set/Packs pairs were moved PAST these columns
    # because they are filled 13 times in 313 rows; these two are filled on
    # every box opening, which is most of the catalogue, so they belong beside
    # Opening Type where the answer is already in mind.
    #
    # BOTH ARE OPTIONAL AND BLANK IS THE NORMAL CASE. 316 rows have neither
    # today. Everything downstream omits what it does not have rather than
    # printing a zero, so a blank row reads exactly as it did before these
    # columns existed.
    ("Box #", 8, "input"),
    ("Pack #", 8, "input"),
    # Packs Opened is what makes the luck page rigorous. Without it a rate can
    # only be "per video", which silently treats a 36-pack booster box and a
    # single pack as one trial each. With it the rate is per PACK, which is the
    # number anyone actually means by "how often do you hit".
    # Now a SUM of the five pack cells, not a typed number. It used to be typed
    # independently, so a row could say 18 packs while its per-set cells added
    # to 12 and nothing would notice. Grey because it is computed.
    ("Packs Opened", 13, "locked"),
    ("Has Hit", 9, "input"),
    ("Hit Card", 30, "input"),
    ("Hit Rarity", 34, "input"),
    # THE FOUR EXTRA SET/PACKS PAIRS LIVE HERE, PAST THE DAILY COLUMNS.
    #
    # They used to sit directly after Set and Packs, which put nine columns he
    # almost never touches between the two halves of his actual work: Set and
    # Packs on one side, Opening Type through Hit Card on the other. Measured on
    # his real sheet, that is about 270 characters of width scrolled past on
    # every one of 313 rows, to reach columns used by 13 rows in total. Set 2 is
    # filled 12 times, Set 3 twice, Set 4 twice, Set 5 never.
    #
    # NOT REMOVED. A UPC spanning five sets genuinely needs somewhere to record
    # the split, and the argument for having them is untouched by where they
    # sit. Each Set keeps its Packs beside it. Nothing reads these by position:
    # import-sheet.mjs looks every column up by header name.
    ("Set 2", 24, "input"),
    ("Packs 2", 8, "input"),
    ("Set 3", 24, "input"),
    ("Packs 3", 8, "input"),
    ("Set 4", 24, "input"),
    ("Packs 4", 8, "input"),
    ("Set 5", 24, "input"),
    ("Packs 5", 8, "input"),
    ("Greatest Hits", 14, "hof"),
    ("Greatest Hits Rank", 18, "hof"),
    ("Playlist To Add", 18, "input"),
    ("Affiliate Link", 30, "input"),
    ("Site Title", 40, "input"),
    ("Short Description", 46, "input"),
    ("Feature", 9, "input"),
    ("Hide", 8, "input"),
    ("Notes", 34, "input"),
]

FILL = {"locked": HEAD_LOCKED, "input": HEAD_INPUT, "hof": HEAD_HOF}

# Hover notes on the headers that need explaining. These survive the trip
# through Google Sheets as cell notes, which is where an instruction actually
# gets read: nobody scrolls back to a Read Me tab mid-row.
HEAD_NOTES = {
    # The rarity column is now a convenience, not a source. It used to be
    # PREFILLED from tags derived from the video title, in blue meaning
    # "confirm this", and the export to Google Sheets throws the colour away, so
    # 62 of 64 guesses came back looking like answers and the site published
    # hits that never happened. It is never prefilled now and it is ignored
    # unless a card is named beside it.
    "Hit Rarity": (
        "OPTIONAL, and it does not drive anything on its own.\n"
        "The rarity the site uses is read out of the Hit Card column.\n"
        "This column is only kept when a card is named beside it, so filling it\n"
        "in on its own does nothing. Use it as a quick label if you like.\n"
        "\n" + RARITY_HINT
    ),
    "Hit Card": (
        "THIS IS THE COLUMN THAT DRIVES THE SITE. Nothing else claims a hit.\n"
        "\n"
        "FREE TEXT. There is no dropdown on this column and there should not be\n"
        "one. If Sheets offers to turn what you have typed into a dropdown, say\n"
        "no; if one appears, select the column and use Data, Data validation,\n"
        "Remove rule. A dropdown here means retyping a card you already wrote.\n"
        "\n"
        "One line per card, in this shape:\n"
        "   Set - Card name - Rarity\n"
        "For example:\n"
        "   Surging Sparks - Feebas - Illustration Rare\n"
        "\n"
        "Several cards from one rip go in the SAME cell, separated by commas,\n"
        "and they can be from different sets and different rarities:\n"
        "   Journey Together - Noibat - Illustration Rare,\n"
        "   Mega Evolution - Marshadow - Illustration Rare\n"
        "\n"
        "The site reads the rarity out of these words, so spell the rarity out.\n"
        "The star description is welcome and ignored: write it if it helps you.\n"
        "\n"
        "Leave it EMPTY when there was no hit. An empty cell means the video\n"
        "shows no rarity badge anywhere on the site, which is correct: we only\n"
        "claim a pull where you have said what the card was."
    ),
    "Packs Opened": (
        "COMPUTED. Do not type here.\n"
        "It adds up Packs, Packs 2, Packs 3, Packs 4 and Packs 5.\n"
        "Put the count for each set in that set's own Packs column and this\n"
        "adds itself up. Typing over it deletes the sum for that row.\n"
        "\n"
        "WHAT THE FILLED COLUMNS BUY, once this sheet comes back:\n"
        "  Opening Type + Box #  ->  exactly how many ETBs, Booster Bundles\n"
        "                            and single packs we have opened, per set\n"
        "                            and overall.\n"
        "  Set + Packs           ->  exactly how many packs of each set we\n"
        "                            have opened, including the ones that came\n"
        "                            out of a mixed box.\n"
        "  Pack #                ->  which pack of that box the video is, so a\n"
        "                            label reads 'Chaos Rising ETB 3 - Pack 3'.\n"
        "Nothing on the site prints any of these until you have typed them."
    ),
    "Box / Series": (
        "FREE TEXT. Type it however you say it out loud.\n"
        "\n"
        "  Pitch Black Booster Bundle #3 Pack#5\n"
        "  Chaos Rising ETB 3 - pack 3\n"
        "  Journey Together Booster Bundle 2, pack 6\n"
        "\n"
        "This one cell can fill Opening Type, Set, Box # and Pack # for you, so\n"
        "if you write it here you do not have to fill those four as well.\n"
        "Anything you DO type in those columns wins over what is read here.\n"
        "\n"
        "There is no dropdown any more, on purpose: the old list could only\n"
        "offer products already in the log, so it could never contain the one\n"
        "you were recording for the first time."
    ),
    "Packs": (
        "HOW MANY PACKS OF THE SET IN THE Set COLUMN you opened in THIS video.\n"
        "Not how many the box holds. If you opened one pack out of a Chaos\n"
        "Rising ETB, this is 1, not 9. The 9 belongs to the box, and Box #\n"
        "and Pack # are where you say which box and which pack.\n"
        "\n"
        "A whole box opened in one video IS the full number: if you opened an\n"
        "entire ETB on camera, put 9 and leave Pack # empty.\n"
        "\n"
        "MIXED VIDEOS USE THE EXTRA PAIRS. A UPC holding 6 Journey Together\n"
        "and 6 Pitch Black packs is Set=Journey Together Packs=6 and\n"
        "Set 2=Pitch Black Packs 2=6. Packs Opened adds them to 12 for you.\n"
        "This is what makes 'how many Pitch Black packs have we opened'\n"
        "answerable at all: a single total cannot be split back apart."
    ),
    "Box #": (
        "WHICH ONE OF THESE, counting your own openings of that product from\n"
        "that set. Your third Chaos Rising ETB is 3. Your second Journey\n"
        "Together Booster Bundle is 2.\n"
        "\n"
        "Counted per SET and per OPENING TYPE, so a Chaos Rising ETB and a\n"
        "Journey Together ETB have their own separate numbering, and so does a\n"
        "bundle from the same set. \"Box\" covers whatever the sealed thing is:\n"
        "an ETB, a bundle, a tin, a collection.\n"
        "\n"
        "It counts boxes you OPENED, on camera or not. If number 1 and 2 were\n"
        "never filmed, this is still 3.\n"
        "\n"
        "Leave it blank when you do not know or it does not apply. A single\n"
        "loose pack has no box, so it stays empty. Blank is normal and nothing\n"
        "on the site shows a gap where one is missing."
    ),
    "Pack #": (
        "WHICH PACK OUT OF THAT BOX this video opens. Pack 3 of the 9 in an\n"
        "ETB is 3.\n"
        "\n"
        "Use it with Box #: \"Box # 3, Pack # 3\" is pack three of your third\n"
        "one. That pair is what lets the site tell one ETB from the next.\n"
        "\n"
        "This is NOT the same as Packs, next door. Packs is how many packs the\n"
        "product holds; this is which single one of them you are opening here.\n"
        "\n"
        "Leave it blank when a video opens a whole product in one go, or when\n"
        "you did not keep count."
    ),
    "Greatest Hits": (
        "The Hall of Fame band on the home page.\n"
        "Separate from the Hits playlist: this is the shortlist, that is a\n"
        "YouTube playlist. A video can be in one, both or neither."
    ),
    "Set": (
        "Start typing and pick from the list. If a set is missing, tell Claude\n"
        "rather than typing it freehand: a name that is not on the list maps to\n"
        "no set id and the import leaves the video untagged.\n"
        "\n"
        "SEALED PRODUCTS THAT ARE NOT A SET. A Trick or Trade bundle, a Victini\n"
        "Illustration Collection, a Mega Heroes mini tin: none of these is an\n"
        "expansion, so there is no set to pick and no guide page to link to.\n"
        "Put 'Not a set (sealed/other)' here and put the product's name in the\n"
        "Box / Series column, which offers all three. Both answers are kept."
    ),
}

for i, (head, width, kind) in enumerate(COLUMNS, start=1):
    c = wv.cell(1, i, head)
    c.font = BOLD
    c.fill = FILL[kind]
    c.border = BOX
    c.alignment = Alignment(vertical="center", wrap_text=True)
    if head in HEAD_NOTES:
        cm = Comment(HEAD_NOTES[head], "Garbage Rips")
        cm.width, cm.height = 320, 190
        c.comment = cm
    wv.column_dimensions[get_column_letter(i)].width = width
# FREEZE THROUGH Watch, NOT JUST THE ID. At B2 only the video id stayed put, so
# scrolling right to the columns he actually fills took the TITLE and the WATCH
# link off screen: he was filling a row he could no longer identify and could no
# longer open. Everything left of the first input column is now pinned.
_HEADS = [h for h, _, _ in COLUMNS]
wv.freeze_panes = f"{get_column_letter(_HEADS.index('Opening Type') + 1)}2"
wv.row_dimensions[1].height = 30

# AUTOFILTER, AND IT IS THE BIGGEST USABILITY FIX IN THIS FILE.
# The log is sorted newest first, which is right for "I just uploaded, log it".
# Tim works the other way for backfill, oldest first, and without a filter that
# meant scrolling to row 314 every session and hunting for the next blank by
# eye. With this he sorts Published ascending in one click, and can filter any
# column to Blanks to see exactly what is left.
# Survives the trip into Google Sheets as its standard filter.
wv.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(videos) + 1}"

# ---------------------------------------------------------------------------
# SHOW WHAT IS STILL MISSING, so a 317 row job can be reviewed at a glance
# ---------------------------------------------------------------------------
#
# Tim, 18 August 2026: "I just dont want to run into anything that slows me
# down". The thing that slows a long log down is not typing, it is FINDING the
# row you have not done yet, and re-checking rows you already finished.
#
# One rule, on the four columns that must be answered for the stats he asked for
# to come out right: a row is unfinished while Opening Type, Set, Packs or Has
# Hit is empty. Amber fill, not red: nothing here is wrong, it is just not
# answered yet, and 317 red rows on the first open would read as 317 errors.
#
# WHY A FORMULA RULE RATHER THAN openpyxl's blanks operator: the blanks operator
# colours a cell when THAT cell is empty, so a finished row with an empty Packs
# still shows three white cells and one amber, which reads as "nearly done"
# rather than "not done". This colours the WHOLE input block while any of the
# four is missing, so an unfinished row is a stripe you cannot miss when
# scrolling, and it clears the moment the row is complete.
#
# GOOGLE SHEETS KEEPS THIS. Conditional formatting survives an .xlsx import,
# unlike the blue "this is a guess" font colour, which is why that convention
# was removed from this file rather than relied on.
_ix = lambda h: _HEADS.index(h) + 1
_first = get_column_letter(min(_ix(h) for h in ("Opening Type", "Set", "Packs", "Has Hit")))
_last = get_column_letter(max(_ix(h) for h in ("Opening Type", "Set", "Packs", "Has Hit")))
_need = "OR({})".format(",".join(
    f'${get_column_letter(_ix(h))}2=""' for h in ("Opening Type", "Set", "Packs", "Has Hit")))
wv.conditional_formatting.add(
    f"{_first}2:{_last}{len(videos) + 1}",
    FormulaRule(formula=[_need], fill=PatternFill("solid", fgColor="FFF3D6"), stopIfTrue=False),
)

# A ROW THAT SAYS THERE WAS A HIT AND DOES NOT NAME THE CARD is the one case
# that IS an error rather than an absence: /luck.html counts it as a hit while
# the hall of fame and the rarity pages have nothing to show. Rose, and only on
# the two cells involved, so it reads as different from "not answered yet".
_hh = get_column_letter(_ix("Has Hit"))
_hc = get_column_letter(_ix("Hit Card"))
wv.conditional_formatting.add(
    f"{_hh}2:{_hc}{len(videos) + 1}",
    FormulaRule(formula=[f'AND(${_hh}2="Yes",${_hc}2="")'],
                fill=PatternFill("solid", fgColor="FADADD"), stopIfTrue=False),
)

COL = {head: i for i, (head, _, _) in enumerate(COLUMNS, start=1)}

# Rows where a stored pack count was the old capacity guess and the video says
# which pack it was. Printed at the end so the correction is never silent.
PACK_CORRECTIONS = []


def clock(sec):
    if not sec:
        return ""
    return f"{sec // 60}:{sec % 60:02d}"

def best_pull(v):
    for p in PULL_ORDER:
        if p in (v.get("pulls") or []):
            return p
    return None

ordered = sorted(videos, key=lambda v: (v.get("published") or ""), reverse=True)

for r, v in enumerate(ordered, start=2):
    vid = v["id"]
    sets_v = v.get("sets") or []
    products = v.get("products") or []
    pull = best_pull(v)

    wv.cell(r, COL["Video ID"], vid).font = LOCKED_TXT
    wv.cell(r, COL["Title"], v.get("title", "")).font = LOCKED_TXT
    wv.cell(r, COL["Published"], v.get("published", "")).font = LOCKED_TXT
    wv.cell(r, COL["Views"], v.get("views", 0)).font = LOCKED_TXT
    wv.cell(r, COL["Length"], clock(v.get("duration"))).font = LOCKED_TXT
    w = wv.cell(r, COL["Watch"], f'=HYPERLINK("https://youtu.be/{vid}","watch")')
    w.font = Font(name="Arial", size=10, color="0563C1", underline="single")

    # Prefilled guesses in blue: correcting beats typing.
    # A video can hold packs from several sets, so spread them across the three
    # columns in order. The first is what the video is really about and picks
    # the wrapper on the site.
    # BY NAME, NEVER BY OFFSET. This used to write the extra sets at
    # COL["Set"] + n*2, on the assumption that Set, Packs, Set 2, Packs 2 sit
    # contiguously. They no longer do: the four extra pairs were moved past the
    # columns Tim fills on every row, and the moment they moved this loop began
    # writing Set 2 into whatever now sits two columns after Set. It wrote
    # nothing at all in practice, so all 12 multi-set rows came back with only
    # their first set and the second was silently dropped on the next import.
    #
    # Looking each column up by header is the same rule import-sheet.mjs already
    # follows, and it is why THAT side survived the move untouched.
    for n, sid in enumerate(sets_v[:5]):
        _key = "Set" if n == 0 else f"Set {n + 1}"
        if sid in set_name and _key in COL:
            wv.cell(r, COL[_key], set_name[sid]).font = GUESS_TXT
    # "NOT A SET" IS AN ANSWER, SO IT HAS TO COME BACK AS ONE. A video whose
    # override says sets:[] has no set on purpose, and reading that back off
    # videos.json gives an empty cell, which is the sheet's word for "nobody has
    # said yet". Handing an answered question back as an unanswered one is how
    # the same row gets filled in three times.
    if not sets_v and overrides_src.get(vid, {}).get("sets") == []:
        wv.cell(r, COL["Set"], "Not a set (sealed/other)").font = BODY
    # OPENING TYPE IS NO LONGER GUESSED EITHER. Tim, 18 August 2026: "yes make
    # the execl sheet let me fill in what type of product it is, then what set
    # type of packs are in it and how many of each of those set type packs are
    # in the overall video".
    #
    # This used to write PRODUCT_TO_OPENING's answer in blue on all 316 rows.
    # The taxonomy reads a title, and a title cannot tell an ex Premium
    # Collection from an ex Box: 24 rows were logged as one and came back as the
    # other, which is recorded a few lines down as the reason the RESTORE for
    # this column had to be added. The same CSV colour loss applies, so the
    # guess became the answer on every row nobody corrected.
    #
    # He is filling this column himself now, and the three columns that depend
    # on it, so a blank cell is the question and nothing pre-answers it.
    # NOTHING IS GUESSED INTO THESE THREE COLUMNS ANY MORE.
    #
    # Tim, 18 August 2026: "make sure you aren't tagging any videos with what
    # type of product it is and what packs are in the video until you get my
    # execl sheet thats filled out with all that exact data".
    #
    # Box #, Pack # and Packs are now handed back EMPTY unless a person has
    # answered them. Earlier today this file prefilled Box # and Pack # from
    # Tim's own titles ("Mega Zygarde Box #2 Pack #5") in blue, and prefilled
    # Packs from PRODUCT_TO_PACKS. Both were suggestions to confirm, and the
    # comment further down explains why a suggestion is not safe here: colour
    # does not survive export to CSV, so the importer reads every blue cell back
    # as a typed answer. That is how "9 packs" ended up on 21 one-pack rips and
    # how "232 packs counted" reached /luck.html.
    #
    # A BLANK CELL IS THE HONEST STATE and it costs nothing downstream: every
    # reader is written as "show it if it is there", so a blank renders nothing
    # rather than a zero. The parse still exists in _stated() and is reported at
    # the end of the run as a convenience, so Tim can see at a glance which rows
    # his own copy already answers, without any of it touching a cell.
    # Packs Opened = the five per-set counts, summed. Typed independently it
    # could say 18 while the parts added to 12 and nothing would catch it.
    _pk = [get_column_letter(COL[h]) for h in ("Packs", "Packs 2", "Packs 3", "Packs 4", "Packs 5")]
    wv.cell(r, COL["Packs Opened"], "=SUM(" + ",".join(f"{c}{r}" for c in _pk) + ")")
    # DO NOT PREFILL A HIT. This used to write "Yes" and a rarity, guessed from
    # words in the title and description, in blue to mean "confirm this". The
    # export to CSV throws the colour away, so the importer read every guess
    # back as a typed answer, and the site then showed a hit badge on videos
    # where the title merely MENTIONED a card being hunted. 62 of 64 rarity
    # values in the log arrived that way and not one of them was typed.
    #
    # A guess that cannot be told apart from an answer downstream is not a
    # convenience, it is a way to publish something nobody said. The set and the
    # opening type are still prefilled, because those describe what was OPENED
    # and a title says that reliably. What was PULLED only the person who
    # opened it knows.

    # WHAT WAS ALREADY IMPORTED COMES BACK, which the Read Me has always claimed
    # and which was true of nothing. This script only ever read videos.json, so
    # a rebuild returned the machine's guesses and dropped every answer a human
    # had given: Has Hit, the rarity, Greatest Hits, the box name, the pack
    # count. Tim corrected a Costco UPC from 16 packs to 18, imported it, and
    # the next rebuild handed back 16 with no indication anything had been said.
    #
    # These win over the guesses above, in BODY rather than GUESS_TXT, because
    # they are answers rather than suggestions.
    man = manual.get(v["id"], {})
    # OPENING TYPE WAS MISSING FROM THIS BLOCK AND THE GUESS ABOVE WAS WINNING.
    # Every other answered field is restored here, so a rebuild handed back the
    # taxonomy's guess over a typed answer and said nothing: 24 videos logged as
    # "ex Premium Collection" came back as "ex Box", because the tag rules cannot
    # tell those two apart from a title and a person can. That is the exact
    # failure the note below describes, in the same block, one field over.
    # OPENING TYPE AND PACKS ARE NOT RESTORED EITHER, AND THIS IS THE HARD ONE.
    #
    # Everything else in this block is a genuine answer coming back so nobody
    # types it twice. These two are different: manual.json cannot tell a value
    # Tim typed from a value this script SUGGESTED and the CSV round trip
    # laundered into an answer. 316 opening types and 244 pack counts were in
    # the file, and the great majority of both arrived as blue guesses.
    #
    # He asked to fill these himself: "yes make the execl sheet let me fill in
    # what type of product it is, then what set type of packs are in it and how
    # many of each of those set type packs are in the overall video". Handing
    # back a laundered guess in BODY, indistinguishable from his own answer, is
    # precisely what stops that from being possible.
    #
    # NOTHING IS DESTROYED. data/manual.json still holds every stored value; it
    # is the sheet, the input form, that starts clean for these two columns. A
    # handful of real corrections he made are in there (a Costco UPC moved from
    # 16 packs to 18) and will need typing once more. That is the cost of not
    # being able to tell them apart, and it is smaller than the cost of shipping
    # 316 machine guesses as his answers a second time.
    # Typed answers, never guessed. There is no rule that could derive which ETB
    # a video opened: the description usually says so in prose and sometimes says
    # it twice with two different numbers ("Pack #2 of our third Chaos Rising
    # ETB"), which is exactly the sort of thing a matcher gets confidently wrong.
    # A blank cell here means nobody has said, and it comes back blank.
    if man.get("boxNumber"):
        wv.cell(r, COL["Box #"], man["boxNumber"]).font = BODY
    if man.get("packNumber"):
        wv.cell(r, COL["Pack #"], man["packNumber"]).font = BODY
    if man.get("hasHit") is not None:
        wv.cell(r, COL["Has Hit"], "Yes" if man["hasHit"] else "No").font = BODY
    if man.get("hitRarity"):
        wv.cell(r, COL["Hit Rarity"], man["hitRarity"]).font = BODY
    if man.get("greatest"):
        wv.cell(r, COL["Greatest Hits"], "Yes").font = BODY
    # hofRank, NOT greatestRank. import-sheet.mjs writes `hofRank` and
    # sync-youtube.mjs reads `hofRank`; this one line was the only place in the
    # project that said greatestRank, so it always read a key that does not
    # exist and every rank a person typed came back blank on the next rebuild.
    # Nine of nine in a round-trip test, silently, in the copy he keeps working
    # in. The Rank column is what orders the gold band at the top of the home
    # page, so it is a column with an obvious job and a 100% loss rate.
    if man.get("hofRank") is not None:
        wv.cell(r, COL["Greatest Hits Rank"], man["hofRank"]).font = BODY
    if man.get("playlistToAdd"):
        wv.cell(r, COL["Playlist To Add"], man["playlistToAdd"]).font = BODY
    if man.get("box"):
        wv.cell(r, COL["Box / Series"], man["box"]).font = BODY
    if man.get("hitCard"):
        wv.cell(r, COL["Hit Card"], man["hitCard"]).font = BODY
    for _k, _c in (("affiliate", "Affiliate Link"), ("siteTitle", "Site Title"),
                   ("blurb", "Short Description"), ("notes", "Notes")):
        if man.get(_k) and _c in COL:
            wv.cell(r, COL[_c], man[_k]).font = BODY
    for _k, _c in (("feature", "Feature"), ("hide", "Hide")):
        if man.get(_k) and _c in COL:
            wv.cell(r, COL[_c], "Yes").font = BODY

    # THE PACK COUNT GOES BACK IN THE FIRST SET'S CELL, and Packs Opened sums
    # to the total that was logged.
    #
    # This first left the cells blank on a multi-set video and put the total in
    # Notes, reasoning that 18 across five sets does not say how many came from
    # each and the sheet should not invent a split. That is true, and it was
    # still the wrong call: it left Tim looking at empty cells where he had
    # typed a number, which reads as data loss no matter what a note says. He
    # knows the split and can move the figure across the columns in seconds;
    # what he cannot do is get back a number the sheet threw away.
    #
    # So the total lands in the first cell and the note says what to do with it.
    # The SUM is right immediately, and the only thing left open is the
    # distribution, which was always his to give.
        for _h in ("Packs 2", "Packs 3", "Packs 4", "Packs 5"):
            if _h in COL:
                wv.cell(r, COL[_h]).value = None

    for col in range(1, len(COLUMNS) + 1):
        cell = wv.cell(r, col)
        cell.border = BOX
        if cell.font is None or cell.font.name != "Arial":
            cell.font = BODY
        if col in (4,):
            cell.alignment = Alignment(horizontal="right")

last = len(ordered) + 1
# Look columns up by header rather than by number: three new columns shifted
# everything after Set, and hand-counted indexes are how a dropdown ends up on
# the wrong column without anything appearing to break.
CI = {head: i for i, (head, _, _) in enumerate(COLUMNS, start=1)}
for dv_formula, cols in [
    (DV_SET, [CI["Set"], CI["Set 2"], CI["Set 3"], CI["Set 4"], CI["Set 5"]]),
    (DV_OPEN, [CI["Opening Type"]]),
    (DV_RARITY, [CI["Hit Rarity"]]),
    (DV_YESNO, [CI["Has Hit"], CI["Greatest Hits"], CI["Feature"], CI["Hide"]]),
    (DV_PLAYLIST, [CI["Playlist To Add"]]),
    (DV_RANK, [CI["Greatest Hits Rank"]]),
    # A list of numbers, not strict, so 3 is one click and anything past the end
    # of the list can still be typed. Own object per column for the reason
    # spelled out below.
    (DV_BOXNO, [CI["Box #"]]),
    (DV_PACKNO, [CI["Pack #"]]),
]:
    # ONE VALIDATION OBJECT PER COLUMN, NOT ONE SHARED ACROSS ALL OF THEM.
    #
    # openpyxl is happy to attach a single DataValidation to several ranges at
    # once, and it writes valid xlsx: one rule with sqref="G2:G312 I2:I312
    # K2:K312 M2:M312 O2:O312". Excel honours that. GOOGLE SHEETS DOES NOT. On
    # import it keeps the rule for the FIRST range and drops the others, then
    # rebuilds a dropdown for the abandoned columns out of whatever values it
    # finds already typed in them.
    #
    # The symptom is horrible to diagnose because the file is correct: Set had
    # all 38 sets and Set 2 offered exactly three, which were simply the three
    # anyone had ever typed into that column. It looked like a truncated list
    # rather than a lost rule.
    #
    # This sheet's whole reason for existing is that it opens in Google Sheets,
    # so it is Google's behaviour that decides what is correct here. Same story
    # for the Yes/No columns, which shared one rule across four.
    for col in cols:
        d = dv(dv_formula)
        wv.add_data_validation(d)
        c = get_column_letter(col)
        d.add(f"{c}2:{c}{last}")

# Suggestion lists: the dropdown helps, but neither list can ever be complete,
# so typing something new must not be rejected.
# HIT CARD HAS NO DROPDOWN, ON PURPOSE. It used to offer the 500 dearest cards
# as a non-strict suggestion list, which sounds helpful and is not, for two
# reasons that both bite in Google Sheets.
#
# The column does not hold ONE card name. A real entry reads "Phantasmal Flames
# - Trainer - Dawn - Double Silver Star - Ultra Rare, Mega Evolution - Mega
# Gardevoir ex - ..." and lists fourteen cards with their sets and rarities. A
# single-select control cannot express that, so the list could never contain a
# valid answer.
#
# And openpyxl's non-strict flag does not survive the trip. Sheets imports it as
# a real dropdown, then rebuilds the option list out of whatever is already in
# the column, so restoring two long hit lists turned into a two-item menu that
# refused anything typed. The field became unfillable by the one person who
# fills it.
#
# Free text. The My Hits tab is where a card gets picked one at a time.
# BOX / SERIES IS FREE TEXT AND THE DROPDOWN IS GONE, at Tim's request on 19
# August 2026: "the box series column is wrong, the drop down isn't showing
# everything it could be. Can you actually just leave that blank not a drop
# down and I will just type in what set or box it is in there, I will just type
# in Pitch Black Booster Bundle #3 Pack#5, that way you know exactly what it is".
#
# HE IS RIGHT THAT THE LIST WAS INCOMPLETE. BOX_NAMES is built from box names
# already SEEN in the log, so it can only ever offer products he has already
# recorded and never the one he is recording for the first time. A dropdown that
# cannot contain the answer is worse than no dropdown: it makes a person hunt a
# list, fail, and then fight the control to type past it.
#
# WHAT HE TYPES THERE IS AN ANSWER, NOT A GUESS, which is the whole reason this
# is safe to read back. The rule this file has been enforcing all day is that a
# number a REGEX worked out from prose is inference and must not be published,
# while a number TIM TYPED after the word Box is his own statement being read
# back. "Pitch Black Booster Bundle #3 Pack#5" is the second kind. So one cell
# he writes once can fill Opening Type, Set, Box # and Pack #, and doing it that
# way is four times less typing on 317 rows.
#
# The parse belongs in import-sheet.mjs, NOT here, and it must never overwrite a
# cell he filled in himself: a typed column always outranks anything read out of
# this one.
_dv_box_removed = True

# =========================================================== 4. Set Notes ===

wn = wb.create_sheet("Set Notes")
NOTE_COLS = [
    ("Set ID", 22, "locked"), ("Set Name", 24, "locked"),
    ("Released", 12, "locked"), ("Cards", 8, "locked"), ("Your rips", 10, "locked"),
    ("Still In Print", 14, "input"),
    # Pack Price USD and Booster Box Price USD used to live here. Removed: the
    # "Ways to open" band on every set page now carries live TCGplayer prices
    # for every sealed product, so a hand-typed price is both duplicated effort
    # and a chance to contradict the live number printed a few inches below it.
    ("Fun Fact 1", 52, "input"), ("Fun Fact 2", 52, "input"),
]
for i, (head, width, kind) in enumerate(NOTE_COLS, start=1):
    c = wn.cell(1, i, head)
    c.font = BOLD; c.fill = FILL[kind]; c.border = BOX
    wn.column_dimensions[get_column_letter(i)].width = width
wn.freeze_panes = "A2"

rip_counts = {}
for v in videos:
    for s in v.get("sets") or []:
        rip_counts[s] = rip_counts.get(s, 0) + 1

for r, s in enumerate(sorted(sets, key=lambda s: s.get("released") or "", reverse=True), start=2):
    wn.cell(r, 1, s["id"]).font = LOCKED_TXT
    wn.cell(r, 2, s["name"]).font = LOCKED_TXT
    wn.cell(r, 3, s.get("released") or "").font = LOCKED_TXT
    wn.cell(r, 4, s.get("total") or s.get("printedTotal") or "").font = LOCKED_TXT
    wn.cell(r, 5, rip_counts.get(s["id"], 0)).font = LOCKED_TXT
    for col in range(1, len(NOTE_COLS) + 1):
        wn.cell(r, col).border = BOX
        if wn.cell(r, col).font.name != "Arial":
            wn.cell(r, col).font = BODY

dv_print = dv(DV_YESNO)
wn.add_data_validation(dv_print)
dv_print.add(f"F2:F{len(sets) + 1}")

wn.cell(len(sets) + 3, 1, "These two facts are not in the card database and are not guessed. "
                          "The set guides omit them until you fill them in.").font = NOTE

# ============================================================= 5. Summary ===

wsum = wb.create_sheet("Summary")
wsum.column_dimensions["A"].width = 34
wsum.column_dimensions["B"].width = 14
wsum.cell(1, 1, "Progress").font = TITLE

L = "'Video Log'"
def CL(head):
    return get_column_letter(COL[head])
metrics = [
    ("Videos in the log", f'=COUNTA({L}!{CL("Video ID")}2:{CL("Video ID")}{last})'),
    ("Set filled in", f'=COUNTA({L}!{CL("Set")}2:{CL("Set")}{last})'),
    ("Still missing a set",
     f'=COUNTA({L}!{CL("Video ID")}2:{CL("Video ID")}{last})-COUNTA({L}!{CL("Set")}2:{CL("Set")}{last})'),
    ("Videos with 2+ sets", f'=COUNTA({L}!{CL("Set 2")}2:{CL("Set 2")}{last})'),
    ("Videos with 4 sets", f'=COUNTA({L}!{CL("Set 4")}2:{CL("Set 4")}{last})'),
    # The rarity column is the one with the most left to do and it was the one
    # the progress tab never counted, so the summary read as further along than
    # the work actually was.
    ("Hit rarity filled in", f'=COUNTA({L}!{CL("Hit Rarity")}2:{CL("Hit Rarity")}{last})'),
    ("Pack count filled in", f'=COUNTA({L}!{CL("Packs")}2:{CL("Packs")}{last})'),
    ("Still missing a pack count",
     f'=COUNTA({L}!{CL("Video ID")}2:{CL("Video ID")}{last})-COUNTA({L}!{CL("Packs")}2:{CL("Packs")}{last})'),
    ("Opening type filled in", f'=COUNTA({L}!{CL("Opening Type")}2:{CL("Opening Type")}{last})'),
    ("Box number filled in", f'=COUNTA({L}!{CL("Box #")}2:{CL("Box #")}{last})'),
    ("Pack number filled in", f'=COUNTA({L}!{CL("Pack #")}2:{CL("Pack #")}{last})'),
    ("Marked as a hit", f'=COUNTIF({L}!{CL("Has Hit")}2:{CL("Has Hit")}{last},"Yes")'),
    ("Hit card named", f'=COUNTA({L}!{CL("Hit Card")}2:{CL("Hit Card")}{last})'),
    ("In Greatest Hits", f'=COUNTIF({L}!{CL("Greatest Hits")}2:{CL("Greatest Hits")}{last},"Yes")'),
    ("Given a rank", f'=COUNTA({L}!{CL("Greatest Hits Rank")}2:{CL("Greatest Hits Rank")}{last})'),
    ("Affiliate links added", f'=COUNTA({L}!{CL("Affiliate Link")}2:{CL("Affiliate Link")}{last})'),
    ("Hidden from the site", f'=COUNTIF({L}!{CL("Hide")}2:{CL("Hide")}{last},"Yes")'),
]
for i, (label, formula) in enumerate(metrics, start=3):
    wsum.cell(i, 1, label).font = BODY
    c = wsum.cell(i, 2, formula)
    c.font = BOLD
    c.alignment = Alignment(horizontal="right")

wsum.cell(len(metrics) + 5, 1,
          "Counts update as you type. Blank cells are skipped by the import, "
          "so a half-filled sheet is fine.").font = NOTE

# ========================================================== 6. Chase Cards ===
#
# Every chase card the set guides show, plus anything on the hunt list. This is
# where PSA 10 prices come from: nothing can fetch them, so this tab is the only
# way they ever reach the site.

wanted_src = {}
try:
    _w = json.loads((ROOT / "data/wanted.json").read_text())
    for c in _w.get("cards", []):
        wanted_src[f"{c['set']}-{c['number']}"] = c
except Exception:
    pass

psa_src = {}
try:
    psa_src = json.loads((ROOT / "data/psa10.json").read_text()).get("prices", {})
except Exception:
    pass

wc = wb.create_sheet("Chase Cards")
CHASE_COLS = [
    ("Key", 22, "locked"), ("Set", 22, "locked"), ("Card", 30, "locked"),
    ("Number", 9, "locked"), ("Rarity", 30, "locked"), ("Raw USD", 11, "locked"),
    ("PSA 10 USD", 13, "input"), ("PSA 10 Checked", 15, "input"),
    ("PSA 10 Source", 24, "input"),
    ("Most Wanted", 13, "hof"), ("Card Hall of Fame", 18, "hof"),
    ("Pulled On", 12, "input"), ("Pulled In Video", 30, "input"),
    ("Why I Want It", 44, "input"),
]
for i, (head, width, kind) in enumerate(CHASE_COLS, start=1):
    c = wc.cell(1, i, head)
    c.font = BOLD; c.fill = FILL[kind]; c.border = BOX
    c.alignment = Alignment(vertical="center", wrap_text=True)
    wc.column_dimensions[get_column_letter(i)].width = width
wc.freeze_panes = "C2"
wc.row_dimensions[1].height = 30

rows_out = []
seen_keys = set()
for st in sorted(sets, key=lambda x: x.get("released") or "", reverse=True):
    for c in (st.get("chase") or []):
        key = f"{st['id']}-{c['number']}"
        seen_keys.add(key)
        rows_out.append((key, st["name"], c["name"], c["number"], c.get("rarity") or "", c.get("price") or None))
# hunt-list cards from sets too new to have chase data yet
for key, w in wanted_src.items():
    if key in seen_keys:
        continue
    st = next((x for x in sets if x["id"] == w["set"]), None)
    rows_out.append((key, st["name"] if st else w["set"], w["name"], w["number"], w.get("rarity") or "", None))

for r, (key, sname, cname, num, rar, raw) in enumerate(rows_out, start=2):
    wc.cell(r, 1, key).font = LOCKED_TXT
    wc.cell(r, 2, sname).font = LOCKED_TXT
    wc.cell(r, 3, cname).font = LOCKED_TXT
    wc.cell(r, 4, num).font = LOCKED_TXT
    wc.cell(r, 5, rar).font = LOCKED_TXT
    cell = wc.cell(r, 6, raw if raw else "")
    cell.font = LOCKED_TXT
    if raw:
        cell.number_format = '$#,##0.00'
    # carry anything already recorded, so re-running never loses a typed price
    prev = psa_src.get(key)
    if isinstance(prev, dict):
        wc.cell(r, 7, prev.get("price")).font = BODY
        wc.cell(r, 8, prev.get("asOf") or "").font = BODY
        wc.cell(r, 9, prev.get("source") or "").font = BODY
    elif isinstance(prev, (int, float)):
        wc.cell(r, 7, prev).font = BODY
    w = wanted_src.get(key)
    if w:
        wc.cell(r, 10, "Yes").font = GUESS_TXT
        wc.cell(r, 11, "Yes" if w.get("got") else "No").font = GUESS_TXT
        if w.get("note"):
            wc.cell(r, 14, w["note"]).font = GUESS_TXT
    for col in range(1, len(CHASE_COLS) + 1):
        wc.cell(r, col).border = BOX
        if wc.cell(r, col).font.name != "Arial":
            wc.cell(r, col).font = BODY
    wc.cell(r, 7).number_format = '$#,##0.00'

# One rule each, for the reason spelled out on the Video Log above: Google
# Sheets keeps only the first range of a multi-range validation on import, so
# Card Hall of Fame would have lost its Yes/No dropdown while Most Wanted kept
# one. That is the column the whole hall.html page is driven from.
for _col, _why in [("J", "Most Wanted"), ("K", "Card Hall of Fame")]:
    dv_c = dv(DV_YESNO)
    wc.add_data_validation(dv_c)
    dv_c.add(f"{_col}2:{_col}{len(rows_out) + 1}")
dv_src = dv(DV_PSASRC, strict=False)
wc.add_data_validation(dv_src)
dv_src.add(f"I2:I{len(rows_out) + 1}")

wc.cell(len(rows_out) + 3, 1,
        "Most Wanted is a card you are still chasing. Card Hall of Fame is one you have actually "
        "pulled, and it appears on /hall.html ranked by value, PSA 10 first and raw where there is "
        "no graded price. A card can be both while you chase a second copy. "
        "Key is what links a row back to a card and must not be edited. PSA 10 Checked is a date "
        "like 2026-08-11: a graded price with no date is not a fact about anything.").font = NOTE

# ============================================================= 6b. My Hits ===
#
# One row per CARD ACTUALLY PULLED, which is the piece the sheet was missing.
#
# The Chase Cards tab next door is a reference list: the good cards in each set,
# whether or not anyone has seen one. This is the opposite, a log of what came
# out of a pack on camera. They are different things and conflating them breaks
# both: a rip can produce several hits, and a hit can be a card nobody had
# listed as a chase card.
#
# Several rows can share one Video ID. That is the point: an ETB opening with
# three hits is three rows.
#
# Video ID is validated against the Video Log so a typo cannot orphan a card,
# and it is the join that puts these cards under the right rip page.

wh = wb.create_sheet("My Hits")
# Only the first three are worth your time. Number, Rarity and Raw NM are looked
# up from the card data on import when they are left blank, because the site
# already knows all 4,481 cards; fill them in only to overrule what it found.
# PSA 10 has no free feed, so that one is genuinely manual.
HIT_COLS = [
    ("Video ID", 14, "input"),
    ("Card", 30, "input"),
    ("Set", 24, "input"),
    ("Number", 9, "locked"),
    ("Rarity", 32, "locked"),
    ("Raw NM USD", 12, "locked"),
    ("PSA 10 USD", 12, "input"),
    ("Hall of Fame", 13, "hof"),
    ("Notes", 40, "input"),
]
for i, (head, width, kind) in enumerate(HIT_COLS, start=1):
    c = wh.cell(1, i, head)
    c.font = BOLD; c.fill = FILL[kind]; c.border = BOX
    c.alignment = Alignment(vertical="center", wrap_text=True)
    wh.column_dimensions[get_column_letter(i)].width = width
wh.freeze_panes = "B2"
wh.row_dimensions[1].height = 30

HI = {head: i for i, (head, _, _) in enumerate(HIT_COLS, start=1)}
HIT_ROWS = 400          # room to grow; blank rows are ignored on import

# Video ID validated against the Video Log's own column, so the join cannot
# break on a typo.
dv_vid = DataValidation(
    type="list",
    formula1=f"='Video Log'!$A$2:$A${len(ordered) + 1}",
    allow_blank=True, showDropDown=False, showErrorMessage=True,
)
wh.add_data_validation(dv_vid)
dv_vid.add(f"A2:A{HIT_ROWS}")

for formula, head, strict in [
    (DV_SET, "Set", False),
    (DV_RARITY, "Rarity", False),
    (DV_YESNO, "Hall of Fame", True),
]:
    d = dv(formula, strict)
    wh.add_data_validation(d)
    col = get_column_letter(HI[head])
    d.add(f"{col}2:{col}{HIT_ROWS}")

for r in range(2, HIT_ROWS + 1):
    for i in range(1, len(HIT_COLS) + 1):
        cell = wh.cell(r, i)
        cell.font = BODY
        cell.border = BOX
    wh.cell(r, HI["Raw NM USD"]).number_format = '"$"#,##0.00'
    wh.cell(r, HI["PSA 10 USD"]).number_format = '"$"#,##0.00'

# PREFILL FROM data/hits.json, because a rebuild used to hand back 400 empty
# rows. Seventeen cards had already been logged by hand across two rips, and
# every one of them would have come back blank in the file you then keep working
# in, which is the worst possible way to lose them: silently, in the copy you
# trust. The tab header above says rebuilding is safe for "anything already
# imported", and this is what makes that true of hits.
#
# Card, Set and Video ID are what a human typed, so they are written back as
# typed. Number, Rarity and Raw NM are the looked-up columns and stay empty
# unless hits.json already carries an override, so the import still fills them
# from the card data rather than freezing today's answer into the sheet.
try:
    _hits = json.loads((ROOT / "data/hits.json").read_text()).get("videos", {})
except Exception:
    _hits = {}
_hrow = 2
for _vid, _cards in _hits.items():
    for _h in _cards:
        if _hrow > HIT_ROWS:
            break
        wh.cell(_hrow, HI["Video ID"], _vid).font = BODY
        wh.cell(_hrow, HI["Card"], _h.get("card", "")).font = BODY
        wh.cell(_hrow, HI["Set"], _h.get("setName") or _h.get("set", "")).font = BODY
        if _h.get("number"):
            wh.cell(_hrow, HI["Number"], _h["number"]).font = BODY
        if _h.get("rarity"):
            wh.cell(_hrow, HI["Rarity"], _h["rarity"]).font = BODY
        # PSA 10 AND NOTES COME BACK TOO, and they did not. Those are the two
        # columns on this tab that nothing can regenerate: there is no free feed
        # for a graded price, and a note is a note. They were written into
        # hits.json by the import and then left out of this write-back, so the
        # rebuilt workbook handed back an empty cell and the next import read
        # that empty cell as "no price". A round trip that deletes the one
        # number only a human can supply is the exact failure this block exists
        # to prevent, one column over.
        if _h.get("psa10"):
            wh.cell(_hrow, HI["PSA 10 USD"], _h["psa10"]).font = BODY
        if _h.get("notes"):
            wh.cell(_hrow, HI["Notes"], _h["notes"]).font = BODY
        if _h.get("hallOfFame"):
            wh.cell(_hrow, HI["Hall of Fame"], "Yes").font = BODY
        _hrow += 1
_prefilled = _hrow - 2

wh.cell(HIT_ROWS + 2, 1, "One row per card pulled. Several rows can share a Video ID.").font = NOTE
wh.cell(HIT_ROWS + 3, 1, "Hall of Fame = Yes marks the all-time best, which get their own page.").font = NOTE

# ================================================================ 7. Shops ===

wsh = wb.create_sheet("Shops")
SHOP_COLS = [("Name", 26, "input"), ("Website", 46, "input"), ("Area", 20, "input"),
             ("What They Are Good For", 34, "input"), ("Blurb", 54, "input"),
             ("Filmed There", 14, "input")]
for i, (head, width, kind) in enumerate(SHOP_COLS, start=1):
    c = wsh.cell(1, i, head)
    c.font = BOLD; c.fill = FILL[kind]; c.border = BOX
    wsh.column_dimensions[get_column_letter(i)].width = width
wsh.freeze_panes = "A2"
try:
    shops_src = json.loads((ROOT / "data/shops.json").read_text()).get("shops", [])
except Exception:
    shops_src = []
for r, sh in enumerate(shops_src, start=2):
    wsh.cell(r, 1, sh.get("name", "")).font = BODY
    wsh.cell(r, 2, sh.get("url", "")).font = BODY
    wsh.cell(r, 3, sh.get("area", "")).font = BODY
    wsh.cell(r, 4, ", ".join(sh.get("goodFor") or [])).font = BODY
    wsh.cell(r, 5, sh.get("blurb", "")).font = BODY
    wsh.cell(r, 6, "Yes" if sh.get("visited") else "No").font = BODY
    for col in range(1, len(SHOP_COLS) + 1):
        wsh.cell(r, col).border = BOX
ROWS = max(len(shops_src) + 1, 40)
dv_s = dv(DV_YESNO)
wsh.add_data_validation(dv_s)
dv_s.add(f"F2:F{ROWS}")
dv_area = dv(DV_AREA, strict=False)
wsh.add_data_validation(dv_area)
dv_area.add(f"C2:C{ROWS}")
dv_good = dv(DV_GOODFOR, strict=False)
wsh.add_data_validation(dv_good)
dv_good.add(f"D2:D{ROWS}")
wsh.cell(max(len(shops_src) + 3, 5), 1,
         "Add a row per shop. Paste the plain page URL: tracking and session parameters are "
         "stripped on build, but a clean link is easier to check.").font = NOTE

# Register the defined names the dropdowns refer to. Done here rather than at
# definition time because every list has to exist on the Lists tab first.
from openpyxl.workbook.defined_name import DefinedName
for _label, _ref in DEFINED.items():
    if _label not in wb.defined_names:
        wb.defined_names.add(DefinedName(_label, attr_text=_ref))

wb.save(OUT)
print(f"  Chase Cards {len(rows_out)} rows")
print(f"  My Hits     empty, {HIT_ROWS - 1} rows ready")
print(f"  Shops       {len(shops_src)} rows")
print(f"Wrote {OUT.relative_to(ROOT)}")
print(f"  My Hits prefilled with {_prefilled} card(s) already logged in data/hits.json")
print(f"  Video Log   {len(ordered)} rows x {len(COLUMNS)} columns")
print(f"  Set Notes   {len(sets)} rows")
print(f"  prefilled:  set {sum(1 for v in ordered if len(v.get('sets') or []) >= 1)}, "
      f"opening {sum(1 for v in ordered if (v.get('products') or [''])[0] in PRODUCT_TO_OPENING)}, "
      f"rarity {sum(1 for v in ordered if (manual.get(v['id']) or {}).get('hitRarity'))} (restored answers only, never guessed)")

# WHAT TIM'S OWN COPY ALREADY ANSWERS, reported and never written into a cell.
# The columns go back blank because a blue guess becomes a typed answer the
# moment the sheet is exported to CSV. This is the same information offered as a
# progress note instead: how many rows he can fill from his own titles rather
# than from memory.
_says_pack = sum(1 for v in ordered if _stated(v, PACK_RE))
_says_box = sum(1 for v in ordered if _stated(v, BOX_RE))
print(f"  Box # / Pack #: handed back BLANK on all {len(ordered)} rows, by request, "
      f"until the filled sheet comes back")
print(f"    for reference only, your own titles or descriptions already state a "
      f"pack number on {_says_pack} rows and a box number on {_says_box}")
